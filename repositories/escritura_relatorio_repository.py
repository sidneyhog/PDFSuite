"""Grava o RelatorioEscrituras. Se `openpyxl` estiver instalado, sai um
.xlsx com abas; senao, um conjunto de .csv (um por aba) numa subpasta.
"""
from __future__ import annotations

import csv
import logging
from pathlib import Path

from models.escritura_relatorio import RelatorioEscrituras
from services.util_faixas import faixas

logger = logging.getLogger("pdfsuite")


def openpyxl_disponivel() -> bool:
    try:
        import openpyxl  # noqa: F401
        return True
    except ImportError:
        return False


class EscrituraRelatorioRepository:
    def __init__(self, reports_dir: Path, folhas_por_livro: int = 400) -> None:
        self._dir = reports_dir
        self._total = folhas_por_livro

    def salvar(self, rel: RelatorioEscrituras, timestamp: str) -> Path:
        self._dir.mkdir(parents=True, exist_ok=True)
        abas = self._montar_abas(rel)
        if openpyxl_disponivel():
            return self._salvar_xlsx(abas, timestamp)
        return self._salvar_csvs(abas, timestamp)

    # ------------------------------------------------------------------ #

    def _montar_abas(self, rel: RelatorioEscrituras) -> "dict[str, tuple[list[str], list[list]]]":
        resumo_h = [
            "Livro", "PastaDiagnostico", "DiagnosticoReal", "Pasta", "FolhasPresentes",
            "FolhasFaltando_qtd", "FolhasFaltando", "Duplicadas_qtd", "Duplicadas_folhas",
            "Anexos", "Conflitos", "Abertura", "Encerramento", "PastaSemFolha", "Avisos",
        ]
        resumo, faltando, dups, confl, anexos, rastreio = [], [], [], [], [], []
        for lv in rel.livros:
            resumo.append([
                lv.numero, lv.diagnostico, lv.diagnostico_real, str(lv.pasta),
                len(lv.folhas_presentes),
                len(lv.folhas_faltando), faixas(lv.folhas_faltando),
                lv.total_duplicadas, faixas(sorted(lv.duplicadas)), lv.total_anexos,
                len(lv.conflitos), "sim" if lv.tem_abertura else "NAO",
                "sim" if lv.tem_encerramento else "NAO",
                faixas(lv.folhas_sem_arquivo), " | ".join(lv.avisos),
            ])
            for n in lv.folhas_faltando:
                rotulo = {1: "termo de abertura", self._total: "termo de encerramento"}.get(n, "")
                faltando.append([lv.numero, lv.diagnostico, n, rotulo])
            for folha, qtd in sorted(lv.duplicadas.items()):
                confusao = lv.pasta / f"{folha:03d}" / "duplicada"
                dups.append([lv.numero, folha, qtd, str(confusao)])
            for folha_lida, livro_cod, origem, pagina, *resto in lv.conflitos:
                situacao = resto[0] if resto else ""
                acao = resto[1] if len(resto) > 1 else ""
                confl.append([lv.numero, folha_lida, livro_cod, situacao, acao, origem, pagina])
            for folha, qtd in sorted(lv.anexos_por_folha.items()):
                anexos.append([lv.numero, folha, qtd])
            for folha in sorted(lv.origem_por_folha):
                rastreio.append([lv.numero, folha, lv.origem_por_folha[folha]])

        abas = {
            "Resumo": (resumo_h, resumo),
            "Folhas Faltando": (["Livro", "Diagnostico", "Folha", "Observacao"], faltando),
            "Duplicadas": (["Livro", "Folha", "CopiasExtras", "Pasta"], dups),
            "Conflitos": (["Livro", "FolhaLida", "LivroDoCodigo", "Situacao", "AcaoSugerida",
                           "Origem", "PaginaOrigem"], confl),
            "Anexos por Folha": (["Livro", "Folha", "QtdAnexos"], anexos),
            "Rastreabilidade": (["Livro", "Folha", "OrigemNoServidor"], rastreio),
        }
        return abas

    def _salvar_xlsx(self, abas, timestamp: str) -> Path:
        import openpyxl
        from openpyxl.styles import Font

        destino = self._dir / f"Relatorio_escrituras_{timestamp}.xlsx"
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        for nome, (headers, rows) in abas.items():
            ws = wb.create_sheet(nome[:31])
            ws.append(headers)
            for c in ws[1]:
                c.font = Font(bold=True)
            for row in rows:
                ws.append(row)
            ws.freeze_panes = "A2"
            if rows:
                ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}{len(rows) + 1}"
            for i, h in enumerate(headers, start=1):
                largura = max(len(str(h)), *(len(str(r[i - 1])) for r in rows[:200])) if rows else len(str(h))
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(largura + 2, 70)
        wb.save(destino)
        logger.info("Relatorio de escrituras salvo em '%s'.", destino)
        return destino

    def _salvar_csvs(self, abas, timestamp: str) -> Path:
        pasta = self._dir / f"Relatorio_escrituras_{timestamp}"
        pasta.mkdir(parents=True, exist_ok=True)
        for nome, (headers, rows) in abas.items():
            arq = pasta / f"{nome.replace(' ', '_')}.csv"
            with open(arq, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(headers)
                w.writerows(rows)
        logger.info("Relatorio de escrituras (CSV) salvo em '%s'.", pasta)
        return pasta
